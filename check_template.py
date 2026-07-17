import openpyxl

wb = openpyxl.load_workbook('Shablon-prais-lista.xlsx', data_only=True)
ws = wb.active
for r in range(1, 3):
    print([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
