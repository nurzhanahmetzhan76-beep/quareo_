import openpyxl
import sys

try:
    print("Loading workbook...")
    wb = openpyxl.load_workbook('product_request_template.xlsx', data_only=True)
    print('Sheets:', wb.sheetnames)
    
    sheet_name = 'ТНВЭД ЕАЭС'
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Top 5 rows from {sheet_name} ---")
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            print(row)
    else:
        print(f"Sheet '{sheet_name}' not found!")
except Exception as e:
    print(f"Error: {e}")
