import openpyxl

wb = openpyxl.load_workbook('ACTIVE.xlsx')
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print("Headers:", headers)
print("Row 2:", [ws.cell(2, c).value for c in range(1, ws.max_column + 1)])
