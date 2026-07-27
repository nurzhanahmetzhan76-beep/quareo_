"""
Kaspi-bot repricing API endpoints.

Full CRUD for repricing rules + manual toggle ON/OFF + logs history.
Auto-sync products from Kaspi Seller API so users don't need to enter SKUs manually.
"""

from __future__ import annotations

import logging
import uuid
import httpx

from fastapi import APIRouter, Depends, HTTPException, Query
import io
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import openpyxl
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from retailpool.database import get_db
from retailpool.models.user import User
from retailpool.models.repricing import RepricingRule, RepricingLog
from retailpool.models.ntin import UserSellerSettings
from retailpool.schemas.repricing import (
    RepricingRuleCreate,
    RepricingRuleUpdate,
    RepricingRuleOut,
    RepricingToggle,
    RepricingLogOut,
)
from retailpool.services.auth_service import get_current_user
from retailpool.services.kaspi_api import KaspiSellerClient

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/repricing", tags=["Repricing Bot"])

MAX_STEP_KZT = 5  # Hard cap — NEVER exceed 5 tenge


# ── Helper ────────────────────────────────────────────────────

async def _get_rule_for_user(
    rule_id: uuid.UUID, user: User, db: AsyncSession
) -> RepricingRule:
    """Fetch a repricing rule and verify ownership."""
    rule = await db.get(RepricingRule, rule_id)
    if not rule or rule.user_id != user.id:
        raise HTTPException(status_code=404, detail="Правило не найдено")
    return rule


# ══════════════════════════════════════════════════════════════
# CRUD — Rules
# ══════════════════════════════════════════════════════════════

@router.post(
    "/rules",
    response_model=RepricingRuleOut,
    status_code=201,
    summary="Create a repricing rule for a product",
)
async def create_rule(
    data: RepricingRuleCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> RepricingRuleOut:
    """Add a product to the repricing bot.

    Bot is OFF by default — user must explicitly toggle it ON.
    Step is capped at 5 KZT maximum.
    """
    # Validate plan for repricing if is_active is true
    user_plan = (current_user.plan or "free").lower()
    if data.is_active and user_plan == "free" and current_user.email != "karimbai.ali10@mail.ru":
        raise HTTPException(
            status_code=403,
            detail="Бот-репрайсер недоступен на бесплатном тарифе. Пожалуйста, обновите тариф."
        )

    # Enforce step cap
    step = min(data.step_kzt, MAX_STEP_KZT)

    # Validate min_price < current_price
    if data.min_price >= data.my_current_price:
        raise HTTPException(
            status_code=400,
            detail="Минимальная цена должна быть ниже текущей цены."
        )

    rule = RepricingRule(
        user_id=current_user.id,
        product_name=data.product_name,
        kaspi_sku=data.kaspi_sku,
        product_url=data.product_url,
        my_merchant_name=data.my_merchant_name,
        my_current_price=data.my_current_price,
        min_price=data.min_price,
        base_price=data.base_price or data.my_current_price,
        step_kzt=step,
        is_active=data.is_active,
        preorder_days=data.preorder_days,
    )

    db.add(rule)
    await db.flush()
    await db.refresh(rule)

    logger.info(
        "Repricing rule created: %s (SKU=%s) by user %s",
        rule.product_name, rule.kaspi_sku, current_user.email,
    )

    return RepricingRuleOut.model_validate(rule)


@router.get(
    "/rules",
    response_model=list[RepricingRuleOut],
    summary="List all repricing rules for the current user",
)
async def list_rules(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[RepricingRuleOut]:
    """Get all products being tracked by the repricing bot."""
    stmt = (
        select(RepricingRule)
        .where(RepricingRule.user_id == current_user.id)
        .order_by(RepricingRule.created_at.desc())
    )
    result = await db.execute(stmt)
    rules = result.scalars().all()
    return [RepricingRuleOut.model_validate(r) for r in rules]


@router.get(
    "/rules/{rule_id}",
    response_model=RepricingRuleOut,
    summary="Get a specific repricing rule",
)
async def get_rule(
    rule_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> RepricingRuleOut:
    rule = await _get_rule_for_user(rule_id, current_user, db)
    return RepricingRuleOut.model_validate(rule)


@router.patch(
    "/rules/{rule_id}",
    response_model=RepricingRuleOut,
    summary="Update a repricing rule",
)
async def update_rule(
    rule_id: uuid.UUID,
    data: RepricingRuleUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> RepricingRuleOut:
    """Update product settings (min_price, step, etc.).

    Step is always capped at 5 KZT.
    """
    rule = await _get_rule_for_user(rule_id, current_user, db)

    update_data = data.model_dump(exclude_unset=True)

    # Enforce step cap
    if "step_kzt" in update_data:
        update_data["step_kzt"] = min(update_data["step_kzt"], MAX_STEP_KZT)

    for field, value in update_data.items():
        setattr(rule, field, value)

    # Validate min < current after update
    if rule.min_price >= rule.my_current_price:
        raise HTTPException(
            status_code=400,
            detail="Минимальная цена должна быть ниже текущей цены."
        )

    await db.flush()
    await db.refresh(rule)
    return RepricingRuleOut.model_validate(rule)


@router.delete(
    "/rules/clear",
    status_code=204,
    summary="Delete all repricing rules for the user",
)
async def clear_all_rules(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> None:
    from sqlalchemy import delete, select
    
    # Find all rule IDs for the user
    rule_ids_stmt = select(RepricingRule.id).where(RepricingRule.user_id == current_user.id)
    
    # Delete all logs associated with those rules first (to prevent foreign key constraint violation)
    logs_stmt = delete(RepricingLog).where(RepricingLog.rule_id.in_(rule_ids_stmt))
    await db.execute(logs_stmt)
    
    # Delete the rules themselves
    stmt = delete(RepricingRule).where(RepricingRule.user_id == current_user.id)
    await db.execute(stmt)
    await db.flush()

@router.delete(
    "/rules/{rule_id}",
    status_code=204,
    summary="Delete a repricing rule",
)
async def delete_rule(
    rule_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Remove a product from the repricing bot (deletes all logs too)."""
    rule = await _get_rule_for_user(rule_id, current_user, db)
    await db.delete(rule)
    await db.flush()


@router.post(
    "/bulk_preorder",
    summary="Bulk update preorder days for all products",
)
async def bulk_preorder(
    body: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Set the same preorder_days for all products of the current user."""
    days = body.get("preorder_days", 0)
    if not isinstance(days, int) or not (0 <= days <= 30):
        raise HTTPException(status_code=400, detail="preorder_days must be between 0 and 30")

    from sqlalchemy import update
    stmt = (
        update(RepricingRule)
        .where(RepricingRule.user_id == current_user.id)
        .values(preorder_days=days)
    )
    result = await db.execute(stmt)
    await db.commit()
    return {"status": "ok", "updated": result.rowcount}


@router.post(
    "/rules/force_run",
    summary="Manually trigger repricing cycle for current user",
)
async def force_run_repricing(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Manually run a repricing cycle for the user's active rules immediately."""
    from retailpool.services.repricing_service import process_single_rule
    
    stmt = select(RepricingRule).where(
        RepricingRule.user_id == current_user.id,
        RepricingRule.is_active == True
    )
    rows = await db.execute(stmt)
    rules = rows.scalars().all()
    
    if not rules:
        return {"status": "ok", "message": "Нет активных товаров для проверки"}
        
    results = []
    for rule in rules:
        rule.owner_telegram_id = current_user.telegram_id
        res = await process_single_rule(rule, None, db)
        results.append(res)
        
    await db.commit()
    return {"status": "ok", "results": results}

# ══════════════════════════════════════════════════════════════
# Toggle ON/OFF — the core user control
# ══════════════════════════════════════════════════════════════


@router.post(
    "/rules/{rule_id}/toggle",
    response_model=RepricingRuleOut,
    summary="Toggle repricing ON or OFF for a product",
)
async def toggle_rule(
    rule_id: uuid.UUID,
    body: RepricingToggle,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> RepricingRuleOut:
    """User explicitly enables or disables the bot for a specific product.

    When toggled OFF, the bot immediately stops monitoring this product.
    """
    user_plan = (current_user.plan or "free").lower()
    if body.is_active and user_plan == "free" and current_user.email != "karimbai.ali10@mail.ru":
        raise HTTPException(
            status_code=403,
            detail="Бот-репрайсер недоступен на бесплатном тарифе. Пожалуйста, обновите тариф."
        )

    rule = await _get_rule_for_user(rule_id, current_user, db)
    old_state = rule.is_active
    rule.is_active = body.is_active

    await db.flush()
    await db.refresh(rule)

    state_str = "ВКЛ" if body.is_active else "ВЫКЛ"
    logger.info(
        "Repricing toggled %s for %s (was %s) by %s",
        state_str, rule.product_name, "ВКЛ" if old_state else "ВЫКЛ",
        current_user.email,
    )

    return RepricingRuleOut.model_validate(rule)


# ══════════════════════════════════════════════════════════════
# Logs — history of price changes
# ══════════════════════════════════════════════════════════════

@router.get(
    "/rules/{rule_id}/logs",
    response_model=list[RepricingLogOut],
    summary="Get price change history for a product",
)
async def get_rule_logs(
    rule_id: uuid.UUID,
    limit: int = Query(default=50, le=200),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[RepricingLogOut]:
    """Get the audit trail of all price changes made by the bot."""
    # Verify ownership
    await _get_rule_for_user(rule_id, current_user, db)

    stmt = (
        select(RepricingLog)
        .where(RepricingLog.rule_id == rule_id)
        .order_by(RepricingLog.created_at.desc())
        .limit(limit)
    )
    result = await db.execute(stmt)
    logs = result.scalars().all()
    return [RepricingLogOut.model_validate(log) for log in logs]


# ══════════════════════════════════════════════════════════════
# Stats — dashboard summary
# ══════════════════════════════════════════════════════════════

@router.get(
    "/stats",
    summary="Get repricing dashboard stats",
)
async def get_stats(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Dashboard KPIs: total products, active bots, total price changes."""
    total_stmt = select(func.count()).select_from(RepricingRule).where(
        RepricingRule.user_id == current_user.id
    )
    active_stmt = select(func.count()).select_from(RepricingRule).where(
        RepricingRule.user_id == current_user.id,
        RepricingRule.is_active == True,  # noqa: E712
    )

    total = (await db.execute(total_stmt)).scalar() or 0
    active = (await db.execute(active_stmt)).scalar() or 0

    # Count total price changes
    rule_ids_stmt = select(RepricingRule.id).where(
        RepricingRule.user_id == current_user.id
    )
    logs_stmt = select(func.count()).select_from(RepricingLog).where(
        RepricingLog.rule_id.in_(rule_ids_stmt)
    )
    total_changes = (await db.execute(logs_stmt)).scalar() or 0

    return {
        "total_products": total,
        "active_bots": active,
        "total_price_changes": total_changes,
    }


# ══════════════════════════════════════════════════════════════
# Sync — auto-import products from Kaspi Seller API
# ══════════════════════════════════════════════════════════════

@router.get(
    "/check-token",
    summary="Check if user has a Kaspi API token configured",
)
async def check_kaspi_token(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Check whether the user has a Kaspi Seller API token saved."""
    stmt = select(UserSellerSettings).where(
        UserSellerSettings.user_id == current_user.id
    )
    result = await db.execute(stmt)
    settings = result.scalar_one_or_none()

    has_token = bool(settings and settings.kaspi_api_key)
    shop_name = settings.kaspi_shop_name if settings else None

    return {
        "has_token": has_token,
        "shop_name": shop_name,
    }


@router.post(
    "/sync",
    summary="Sync products from Kaspi Seller API",
)
async def sync_products(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Fetch all products from Kaspi XML Feed and create repricing rules.

    Products that already exist (by kaspi_sku) are skipped.
    Bot is OFF by default for all synced products.
    """
    # 1. Get user's Kaspi settings
    stmt = select(UserSellerSettings).where(
        UserSellerSettings.user_id == current_user.id
    )
    result = await db.execute(stmt)
    settings = result.scalar_one_or_none()

    if not settings or not settings.kaspi_xml_url:
        raise HTTPException(
            status_code=400,
            detail="Сначала добавьте ссылку на Kaspi XML прайс-лист в настройках NTIN."
        )

    # 2. Fetch Kaspi XML Feed
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(settings.kaspi_xml_url)
            resp.raise_for_status()
            xml_text = resp.text
    except Exception as e:
        logger.exception("XML download failed: %s", e)
        raise HTTPException(
            status_code=502,
            detail="Не удалось скачать XML прайс-лист Kaspi."
        )

    # 3. Parse XML
    from xml.etree import ElementTree as ET
    try:
        root = ET.fromstring(xml_text)
        # Strip namespaces
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]
        offers = root.findall('.//offer')
    except Exception as e:
        logger.exception("XML parse failed: %s", e)
        raise HTTPException(status_code=500, detail="Ошибка парсинга XML.")

    if not offers:
        return {"synced": 0, "skipped": 0, "message": "В XML файле нет товаров (offer)."}

    # 4. Get existing SKUs to avoid duplicates
    existing_stmt = select(RepricingRule.kaspi_sku).where(
        RepricingRule.user_id == current_user.id
    )
    existing_result = await db.execute(existing_stmt)
    existing_skus = {row[0] for row in existing_result.all()}

    # 5. Create rules for new products
    synced = 0
    skipped = 0
    merchant_name = settings.kaspi_shop_name

    for offer in offers:
        sku = offer.get('sku')
        if not sku:
            continue

        if str(sku) in existing_skus:
            skipped += 1
            continue

        model_node = offer.find('.//model')
        name = model_node.text.strip() if model_node is not None and model_node.text else f"Товар {sku}"
        
        price_node = offer.find('.//price')
        try:
            price = float(price_node.text) if price_node is not None and price_node.text else 0
        except ValueError:
            price = 0

        rule = RepricingRule(
            user_id=current_user.id,
            product_name=str(name)[:512],
            kaspi_sku=str(sku),
            product_url=None,  # user fills this later
            my_merchant_name=merchant_name,
            my_current_price=price,
            min_price=price * 0.9 if price else 0,  # default: 90% of current
            base_price=price,
            step_kzt=5,
            is_active=False,  # OFF by default!
        )
        db.add(rule)
        synced += 1

    await db.flush()

    logger.info(
        "Kaspi sync for %s: synced=%d, skipped=%d",
        current_user.email, synced, skipped,
    )

    return {
        "synced": synced,
        "skipped": skipped,
        "message": f"Импортировано {synced} товаров. Пропущено (уже есть): {skipped}.",
    }
    

@router.post(
    "/upload_sync",
    summary="Sync products from Kaspi Excel/XML File",
)
async def upload_sync(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Fetch all products from uploaded Kaspi Excel or XML file and create repricing rules."""
    stmt = select(UserSellerSettings).where(
        UserSellerSettings.user_id == current_user.id
    )
    result = await db.execute(stmt)
    settings = result.scalar_one_or_none()
    merchant_name = settings.kaspi_shop_name if settings else None

    raw = await file.read()
    filename = file.filename.lower() if file.filename else ""

    # Get existing SKUs
    existing_stmt = select(RepricingRule.kaspi_sku).where(
        RepricingRule.user_id == current_user.id
    )
    existing_result = await db.execute(existing_stmt)
    existing_skus = {row[0] for row in existing_result.all()}

    synced = 0
    skipped = 0

    if filename.endswith(".xml") or raw.startswith(b"<?xml"):
        # PARSE XML FILE
        import xml.etree.ElementTree as ET
        try:
            root = ET.fromstring(raw.decode('utf-8'))
            # Strip namespaces
            for elem in root.iter():
                if '}' in elem.tag:
                    elem.tag = elem.tag.split('}', 1)[1]
            offers = root.findall('.//offer')
            
            for offer in offers:
                sku = offer.get('sku') or offer.get('id')
                if not sku:
                    continue
                if str(sku) in existing_skus:
                    skipped += 1
                    continue
                existing_skus.add(str(sku))
                
                title_elem = offer.find('model')
                title = title_elem.text if title_elem is not None else f"Товар {sku}"
                
                price_elem = offer.find('price')
                try:
                    price = float(price_elem.text) if price_elem is not None else 0
                except (ValueError, TypeError):
                    price = 0

                rule = RepricingRule(
                    user_id=current_user.id,
                    product_name=str(title)[:255],
                    kaspi_sku=str(sku),
                    my_merchant_name=merchant_name,
                    my_current_price=price,
                    min_price=price * 0.9 if price else 0,
                    base_price=price,
                    step_kzt=5,
                    is_active=False,
                )
                db.add(rule)
                synced += 1
                
                if synced % 20 == 0:
                    try:
                        await db.flush()
                    except Exception as e:
                        logger.exception("Database batch flush error in upload_sync (XML): %s", e)
                        raise HTTPException(status_code=400, detail=f"Ошибка базы данных (XML batch): {repr(e)}")

        except Exception as e:
            logger.exception("Failed to parse Kaspi XML file: %s", e)
            raise HTTPException(status_code=400, detail=f"Ошибка парсинга XML файла: {repr(e)}")

    else:
        # PARSE EXCEL FILE
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw))
        except Exception:
            raise HTTPException(status_code=400, detail="Не удалось открыть файл. Загрузите Excel или XML.")

        try:
            ws = wb.worksheets[0]
            headers = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(1, c).value
                if v is not None:
                    headers[str(v).strip().lower()] = c

            if not headers:
                raise HTTPException(status_code=400, detail="Файл пуст или не содержит заголовков.")

            sku_col = None
            price_col = None
            for h_key, h_col in headers.items():
                if any(p in h_key for p in ["sku", "артикул", "код"]):
                    sku_col = h_col
                if any(p in h_key for p in ["price", "цена"]):
                    price_col = h_col

            if not sku_col or not price_col:
                raise HTTPException(
                    status_code=400,
                    detail="Неверный формат Excel: нет колонок SKU (Артикул) или Цена."
                )
            
            name_col = None
            # Look for partial matches to catch 'название товара', 'наименование', etc.
            for h_key, h_col in headers.items():
                if any(p in h_key for p in ["model", "name", "название", "наименовани", "title"]):
                    name_col = h_col
                    break
            
            if not name_col:
                name_col = sku_col

            for r in range(2, ws.max_row + 1):
                sku = ws.cell(r, sku_col).value
                if sku is None:
                    continue
                sku = str(sku).strip()
                
                if sku in existing_skus:
                    skipped += 1
                    continue
                existing_skus.add(sku)

                name = ws.cell(r, name_col).value
                name = str(name).strip() if name else f"Товар {sku}"
                
                price_val = ws.cell(r, price_col).value
                try:
                    price = float(price_val) if price_val is not None else 0
                except (ValueError, TypeError):
                    price = 0

                rule = RepricingRule(
                    user_id=current_user.id,
                    product_name=str(name)[:255],
                    kaspi_sku=str(sku),
                    my_merchant_name=merchant_name,
                    my_current_price=price,
                    min_price=price * 0.9 if price else 0,
                    base_price=price,
                    step_kzt=5,
                    is_active=False,
                )
                db.add(rule)
                synced += 1
                
                if synced % 20 == 0:
                    try:
                        await db.flush()
                    except Exception as e:
                        logger.exception("Database batch flush error in upload_sync: %s", e)
                        raise HTTPException(status_code=400, detail=f"Ошибка БД при пакетном сохранении (Excel): {repr(e)}")
        except HTTPException:
            raise
        except Exception as e:
            logger.exception("Unexpected error processing Excel in upload_sync: %s", e)
            raise HTTPException(status_code=400, detail=f"Неожиданная ошибка обработки Excel: {repr(e)}")

    try:
        await db.flush()
    except Exception as e:
        logger.exception("Database flush error in upload_sync: %s", e)
        raise HTTPException(status_code=400, detail=f"Ошибка базы данных при сохранении (итог): {repr(e)}")

    return {
        "synced": synced,
        "skipped": skipped,
        "message": f"Импортировано {synced} товаров. Пропущено (уже есть): {skipped}.",
    }


from typing import Optional

@router.post(
    "/upload_assortment",
    summary="Onboarding: Единая загрузка Активных и Архивных товаров",
)
async def upload_assortment(
    active_file: Optional[UploadFile] = File(None),
    archive_file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Точка входа (Single Source of Truth) для новых селлеров.
    Читает ACTIVE.xlsx и ARCHIVE.xlsx одновременно, объединяя их в БД.
    Активные товары получают is_active=True, архивные is_active=False.
    """
    if not active_file and not archive_file:
        raise HTTPException(status_code=400, detail="Необходимо загрузить хотя бы один файл (Актив или Архив).")

    stmt = select(UserSellerSettings).where(UserSellerSettings.user_id == current_user.id)
    result = await db.execute(stmt)
    settings = result.scalar_one_or_none()
    merchant_name = settings.kaspi_shop_name if settings else None

    # Загружаем существующие SKU для защиты от дублей
    existing_stmt = select(RepricingRule.kaspi_sku).where(RepricingRule.user_id == current_user.id)
    existing_result = await db.execute(existing_stmt)
    existing_skus = {row[0] for row in existing_result.all()}

    stats = {"synced_active": 0, "synced_archive": 0, "skipped": 0}

    async def process_excel_file(file: UploadFile, is_active: bool, stats_key: str):
        raw = await file.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw))
        except Exception:
            raise HTTPException(status_code=400, detail=f"Не удалось открыть файл {file.filename}. Проверьте формат.")

        ws = wb.worksheets[0]
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None:
                headers[str(v).strip().lower()] = c

        sku_col = next((h_col for h_key, h_col in headers.items() if any(p in h_key for p in ["sku", "артикул", "код"])), None)
        price_col = next((h_col for h_key, h_col in headers.items() if any(p in h_key for p in ["price", "цена"])), None)
        name_col = next((h_col for h_key, h_col in headers.items() if any(p in h_key for p in ["model", "name", "название", "наименовани", "title"])), sku_col)

        if not sku_col or not price_col:
            raise HTTPException(
                status_code=400,
                detail=f"В файле {file.filename} нет обязательных колонок (SKU или Цена)."
            )

        for r in range(2, ws.max_row + 1):
            sku = ws.cell(r, sku_col).value
            if sku is None:
                continue
            
            sku_str = str(int(sku)) if isinstance(sku, float) else str(sku).strip()
            
            # Если товар уже загружен из другого файла или был в базе - пропускаем
            if sku_str in existing_skus:
                stats['skipped'] += 1
                continue
                
            existing_skus.add(sku_str)

            name_val = ws.cell(r, name_col).value if name_col else None
            name = str(name_val)[:255] if name_val is not None else f"Товар {sku_str}"
            
            price_val = ws.cell(r, price_col).value
            try:
                price = float(price_val) if price_val is not None else 0
            except (ValueError, TypeError):
                price = 0

            rule = RepricingRule(
                user_id=current_user.id,
                product_name=name,
                kaspi_sku=sku_str,
                my_merchant_name=merchant_name,
                my_current_price=price,
                min_price=price * 0.9 if price else 0, # По умолчанию дно -10% от текущей
                base_price=price,
                step_kzt=5,
                is_active=is_active, # Вот тут творится магия
            )
            db.add(rule)
            stats[stats_key] += 1
            
            if (stats['synced_active'] + stats['synced_archive']) % 50 == 0:
                await db.flush()

    if active_file:
        await process_excel_file(active_file, is_active=True, stats_key='synced_active')
        
    if archive_file:
        await process_excel_file(archive_file, is_active=False, stats_key='synced_archive')

    try:
        await db.commit()
    except Exception as e:
        logger.exception("Database commit error in upload_assortment: %s", e)
        raise HTTPException(status_code=500, detail="Ошибка при сохранении ассортимента в базу данных.")

    return {
        "synced_active": stats['synced_active'],
        "synced_archive": stats['synced_archive'],
        "skipped": stats['skipped'],
        "message": f"Успешно! Добавлено в Актив: {stats['synced_active']}, в Архив: {stats['synced_archive']}. Пропущено дублей: {stats['skipped']}."
    }


@router.post(
    "/sync_from_ntin",
    summary="Sync products from NTIN database",
)
async def sync_from_ntin(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Import products from the user's NTIN database into Repricing."""
    from retailpool.models.ntin import NtinProduct

    stmt = select(UserSellerSettings).where(
        UserSellerSettings.user_id == current_user.id
    )
    result = await db.execute(stmt)
    settings = result.scalar_one_or_none()
    merchant_name = settings.kaspi_shop_name if settings else None

    # Get all NTIN products for user
    ntin_stmt = select(NtinProduct).where(NtinProduct.user_id == current_user.id)
    ntin_res = await db.execute(ntin_stmt)
    ntin_products = ntin_res.scalars().all()

    if not ntin_products:
        raise HTTPException(
            status_code=400,
            detail="В вашей базе NTIN нет товаров. Добавьте товары в разделе NTIN."
        )

    # Get existing SKUs
    existing_stmt = select(RepricingRule.kaspi_sku).where(
        RepricingRule.user_id == current_user.id
    )
    existing_result = await db.execute(existing_stmt)
    existing_skus = {row[0] for row in existing_result.all()}

    synced = 0
    skipped = 0

    for product in ntin_products:
        sku = product.kaspi_sku
        if not sku:
            continue
            
        if sku in existing_skus:
            skipped += 1
            continue

        price = product.price if product.price else 0

        rule = RepricingRule(
            user_id=current_user.id,
            product_name=product.title_ru[:512],
            kaspi_sku=str(sku),
            product_url=None,
            my_merchant_name=merchant_name,
            my_current_price=price,
            min_price=price * 0.9 if price else 0,
            base_price=price,
            step_kzt=5,
            is_active=False,
        )
        db.add(rule)
        synced += 1

    await db.flush()

    return {
        "synced": synced,
        "skipped": skipped,
        "message": f"Импортировано из NTIN: {synced} товаров. Пропущено (уже есть): {skipped}."
    }

@router.post("/process_excel", summary="Обновить цены в прайс-листе Kaspi")
async def process_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Принимает ПОЛНЫЙ Excel-прайс или XML-файл пользователя из кабинета Kaspi.
    Обновляет ТОЛЬКО колонку price у товаров, где бот включён (is_active),
    беря новую цену из my_current_price. Сохраняет остатки.
    Возвращает готовый Excel-файл для загрузки в Kaspi.
    """
    result = await db.execute(
        select(RepricingRule).where(
            RepricingRule.user_id == current_user.id
        )
    )
    rules = result.scalars().all()
    
    # Store the full rule object instead of just the price
    rule_map = {
        str(r.kaspi_sku).strip(): r
        for r in rules
        if r.kaspi_sku
    }

    if not rule_map:
        raise HTTPException(
            status_code=400,
            detail="Нет активных товаров с посчитанной ценой. Включите бота и дождитесь расчёта.",
        )

    raw = await file.read()
    filename = file.filename.lower() if file.filename else ""
    changed = 0

    if filename.endswith(".xml") or raw.startswith(b"<?xml"):
        import xml.etree.ElementTree as ET
        try:
            root = ET.fromstring(raw.decode('utf-8'))
            for elem in root.iter():
                if '}' in elem.tag:
                    elem.tag = elem.tag.split('}', 1)[1]
            offers = root.findall('.//offer')
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Лист1"
            headers = ['SKU', 'model', 'brand', 'price', 'PP1', 'PP2', 'PP3', 'PP4', 'PP5', 'preOrder']
            ws.append(headers)

            for offer in offers:
                sku = offer.get('sku') or offer.get('id')
                if not sku:
                    continue
                
                sku_str = str(sku).strip()
                
                model_node = offer.find('model')
                model = model_node.text if model_node is not None and model_node.text else f"Товар {sku_str}"
                
                brand_node = offer.find('brand')
                brand = brand_node.text if brand_node is not None and brand_node.text else "Без бренда"
                
                # Default price from XML
                price = ""
                price_node = offer.find('price')
                if price_node is not None and price_node.text:
                    price = price_node.text
                else:
                    cityprice = offer.find('.//cityprice')
                    if cityprice is not None and cityprice.text:
                        price = cityprice.text
                try:
                    price = float(price) if price else 0
                except ValueError:
                    price = 0

                preorder_days = ""
                # OVERRIDE PRICE and PREORDER if bot is active for this SKU
                if sku_str in rule_map:
                    rule = rule_map[sku_str]
                    if rule.is_active and rule.my_current_price:
                        price = rule.my_current_price
                    if rule.preorder_days and rule.preorder_days > 0:
                        preorder_days = min(rule.preorder_days, 30)
                    changed += 1

                # Quantities
                pp_stocks = {'PP1': 'no', 'PP2': 'no', 'PP3': 'no', 'PP4': 'no', 'PP5': 'no'}
                availabilities = offer.findall('.//availability')
                for av in availabilities:
                    store_id = av.get('storeId', '')
                    avail_attr = av.get('available', 'no').lower()
                    stock_val = av.get('stockCount')
                    
                    val_to_write = avail_attr
                    if stock_val is not None:
                        try:
                            val_to_write = int(float(stock_val))
                        except ValueError:
                            pass
                    
                    pp_key = None
                    if 'PP1' in store_id: pp_key = 'PP1'
                    elif 'PP2' in store_id: pp_key = 'PP2'
                    elif 'PP3' in store_id: pp_key = 'PP3'
                    elif 'PP4' in store_id: pp_key = 'PP4'
                    elif 'PP5' in store_id: pp_key = 'PP5'
                    elif 'PP6' in store_id: pp_key = 'PP1'
                    else: pp_key = 'PP1'
                    
                    if pp_key:
                        existing = pp_stocks[pp_key]
                        if isinstance(existing, int) and isinstance(val_to_write, int):
                            pp_stocks[pp_key] = existing + val_to_write
                        elif val_to_write != 'no':
                            pp_stocks[pp_key] = val_to_write

                row = [
                    sku_str, model, brand, price,
                    pp_stocks['PP1'], pp_stocks['PP2'], pp_stocks['PP3'], 
                    pp_stocks['PP4'], pp_stocks['PP5'], preorder_days
                ]
                ws.append(row)
                
        except Exception as e:
            logger.exception("Failed to parse Kaspi XML file in process_excel: %s", e)
            raise HTTPException(status_code=400, detail=f"Ошибка парсинга XML файла: {str(e)}")

    else:
        # Excel processing
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw))
        except Exception:
            raise HTTPException(status_code=400, detail="Не удалось открыть файл. Нужен Excel или XML.")

        ws = wb.worksheets[0]
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None:
                headers[str(v).strip().lower()] = c

        if not headers:
            raise HTTPException(status_code=400, detail="Файл пуст или не содержит заголовков.")

        sku_col = None
        price_col = None
        for h_key, h_col in headers.items():
            if any(p in h_key for p in ["sku", "артикул", "код"]):
                sku_col = h_col
            if any(p in h_key for p in ["price", "цена"]):
                price_col = h_col

        if not sku_col or not price_col:
            raise HTTPException(
                status_code=400,
                detail="Неверный формат: в файле нет колонок SKU (Артикул) или Цена.",
            )
        
        # Look for existing preorder column by partial match
        preorder_col = None
        for h_key, h_col in headers.items():
            if any(p in h_key.lower() for p in ["предзаказ", "preorder"]):
                preorder_col = h_col
                break
        
        # Get PP columns to activate archived products
        pp_cols = [headers[k] for k in headers if k.startswith("pp")]

        for r in range(2, ws.max_row + 1):
            sku = ws.cell(r, sku_col).value
            if sku is None:
                continue
                
            # ensure string, handle .0 floats
            if isinstance(sku, float):
                sku = str(int(sku))
            else:
                sku = str(sku).strip()
                
            matched_rule = None
            if sku in rule_map:
                matched_rule = rule_map[sku]
            else:
                # Handle Kaspi composite SKUs like "118285743_372127156"
                for part in sku.split('_'):
                    if part in rule_map:
                        matched_rule = rule_map[part]
                        break
                        
            if matched_rule:
                if matched_rule.is_active and matched_rule.my_current_price:
                    ws.cell(r, price_col).value = matched_rule.my_current_price
                
                if matched_rule.preorder_days and matched_rule.preorder_days > 0:
                    preorder_val = min(matched_rule.preorder_days, 30)
                    if preorder_col:
                        ws.cell(r, preorder_col).value = preorder_val
                    else:
                        # Append the column if it doesn't exist
                        preorder_col = ws.max_column + 1
                        ws.cell(1, preorder_col).value = "preOrder"
                        headers["preorder"] = preorder_col
                        ws.cell(r, preorder_col).value = preorder_val
                        
                    # Force availability to 'yes' for preorders so archived items become active
                    for col_idx in pp_cols:
                        val = ws.cell(r, col_idx).value
                        if not val or str(val).strip().lower() == "no":
                            ws.cell(r, col_idx).value = "yes"

                changed += 1

    logger.info("process_excel: user=%s changed=%d rows", current_user.id, changed)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    # Send back as Excel
    download_name = file.filename or "kaspi_updated_prices"
    if download_name.endswith(".xml"):
        download_name = download_name[:-4] + ".xlsx"
    elif not download_name.endswith(".xlsx"):
        download_name += ".xlsx"
        
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'}
    )

@router.post("/generate_base_xml_from_excel", summary="Создать базовый XML из Excel для автоматизации")
async def generate_base_xml_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Parses a Kaspi Excel file and generates a static Kaspi XML catalog.
    Saves it locally and links it in UserSellerSettings so the user can use the automated feed feature
    without needing an external 1C/MoiSklad system.
    """
    import os
    import xml.etree.ElementTree as ET
    
    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw))
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось открыть файл. Загрузите Excel-прайс из Kaspi.")

    ws = wb.worksheets[0]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            headers[str(v).strip().lower()] = c

    sku_col = None
    price_col = None
    for h_key, h_col in headers.items():
        if any(p in h_key for p in ["sku", "артикул", "код"]):
            sku_col = h_col
        if any(p in h_key for p in ["price", "цена"]):
            price_col = h_col

    if not sku_col or not price_col:
        raise HTTPException(status_code=400, detail="В файле нет колонок SKU (Артикул) или Цена.")
        
    name_col = None
    for h_key, h_col in headers.items():
        if any(p in h_key for p in ["model", "name", "название", "наименовани", "title"]):
            name_col = h_col
            break
    
    brand_col = None
    for h_key, h_col in headers.items():
        if "brand" in h_key or "бренд" in h_key:
            brand_col = h_col
            break

    pp_cols = {k: headers[k] for k in headers if any(f"pp{i}" in k for i in range(1, 10))}

    root = ET.Element("kaspi_catalog")
    root.set("date", "2024-01-01 00:00")
    root.set("xmlns", "kaspiShopping")
    root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
    root.set("xsi:schemaLocation", "kaspiShopping http://kaspi.kz/kaspishopping.xsd")
    
    company = ET.SubElement(root, "company")
    company.text = "Quareo Merchant"
    merchantid = ET.SubElement(root, "merchantid")
    merchantid.text = "1234567"
    
    offers = ET.SubElement(root, "offers")

    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, sku_col).value
        if sku is None:
            continue
        if isinstance(sku, float): sku = str(int(sku))
        else: sku = str(sku).strip()
            
        price_val = ws.cell(r, price_col).value
        try: price = float(price_val) if price_val is not None else 0
        except (ValueError, TypeError): price = 0
            
        name_val = ws.cell(r, name_col).value if name_col else None
        brand_val = ws.cell(r, brand_col).value if brand_col else None
        
        name = str(name_val) if name_val is not None else f"Товар {sku}"
        
        offer = ET.SubElement(offers, "offer", sku=sku)
        ET.SubElement(offer, "model").text = name
        
        brand_elem = ET.SubElement(offer, "brand")
        if brand_val is not None:
            brand_elem.text = str(brand_val)
            
        ET.SubElement(offer, "price").text = str(int(price))
        
        availabilities = ET.SubElement(offer, "availabilities")
        has_pp = False
        import re
        for pp_name, col_idx in pp_cols.items():
            has_pp = True
            val = ws.cell(r, col_idx).value
            
            # Extract just "PP1", "PP2", etc from the column name
            match = re.search(r'(pp\d+)', pp_name, re.IGNORECASE)
            store_id = match.group(1).upper() if match else "PP1"
            
            is_avail = False
            stock_count_str = None
            if val is not None:
                val_str = str(val).strip().lower()
                if val_str in ['yes', 'да', 'true', '+']:
                    is_avail = True
                elif val_str in ['no', 'нет', 'false', '-']:
                    is_avail = False
                else:
                    try:
                        f_val = float(val_str)
                        if f_val > 0:
                            is_avail = True
                            stock_count_str = str(f_val)
                        else:
                            is_avail = False
                            stock_count_str = "0.0"
                    except ValueError:
                        pass
            
            attrs = {"available": "yes" if is_avail else "no", "storeId": store_id}
            if stock_count_str is not None:
                attrs["stockCount"] = stock_count_str
                
            ET.SubElement(availabilities, "availability", **attrs)
                
        if not has_pp:
            ET.SubElement(availabilities, "availability", available="no", storeId="PP1")

    xml_str = ET.tostring(root, encoding="utf-8", xml_declaration=True).decode("utf-8")
    
    # Save the file to frontend/assets/feeds/
    # Ensure directory exists
    feeds_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "frontend", "assets", "feeds")
    os.makedirs(feeds_dir, exist_ok=True)
    
    file_path = os.path.join(feeds_dir, f"{current_user.id}.xml")
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(xml_str)
        
    # Update user settings
    # URL will be accessible via Quareo static assets mount
    # Using public domain for production compatibility.
    local_url = f"https://quareo.pro/assets/feeds/{current_user.id}.xml"
    
    stmt = select(UserSellerSettings).where(UserSellerSettings.user_id == current_user.id)
    result = await db.execute(stmt)
    settings = result.scalar_one_or_none()
    
    if settings:
        settings.kaspi_xml_url = local_url
    else:
        new_settings = UserSellerSettings(user_id=current_user.id, kaspi_xml_url=local_url)
        db.add(new_settings)
        
    await db.commit()
    
    return {"status": "ok", "message": "Базовый XML фид успешно сгенерирован и привязан к профилю!", "url": local_url}
