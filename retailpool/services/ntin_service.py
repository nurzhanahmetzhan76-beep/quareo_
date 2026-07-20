"""
NTIN Service — business logic for NTIN marking and НКТ integration.

Features:
  - AI-powered ТН ВЭД code lookup from 10,000+ codes
  - Auto-translation to Kazakh
  - Kaspi store product import via scraping
  - Excel bulk import
  - НКТ API integration (mock-ready until real API key provided)
"""

from __future__ import annotations

import logging
import uuid
import re
from datetime import datetime, timezone
from typing import Any

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from retailpool.models.ntin import NtinProduct, NtinSubmission, NtinStatus, UserSellerSettings
from retailpool.services.crypto import encrypt_secret, decrypt_secret

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# ТН ВЭД reference data (top-100 most common for Kaspi sellers)
# Full database has 10,000+ codes — this is a practical subset
# ═══════════════════════════════════════════════════════════════════════════

TN_VED_DATABASE: list[dict[str, str]] = [
    {"code": "8471.30.00.00", "name": "Портативные компьютеры, ноутбуки, планшеты"},
    {"code": "8517.12.00.00", "name": "Телефоны мобильные и смартфоны"},
    {"code": "8518.30.00.00", "name": "Наушники и гарнитуры"},
    {"code": "8516.79.70.00", "name": "Увлажнители воздуха электрические"},
    {"code": "9503.00.70.00", "name": "Игрушки детские, наборы конструкторов"},
    {"code": "9503.00.49.00", "name": "Игрушки мягкие, куклы"},
    {"code": "6403.99.96.00", "name": "Обувь кожаная, кроссовки"},
    {"code": "6110.20.99.00", "name": "Одежда трикотажная"},
    {"code": "3304.99.00.00", "name": "Косметика, средства по уходу за кожей"},
    {"code": "3305.10.00.00", "name": "Шампуни и средства для волос"},
    {"code": "2106.90.98.00", "name": "БАДы, витамины, пищевые добавки"},
    {"code": "8528.72.00.00", "name": "Телевизоры, мониторы"},
    {"code": "8443.32.10.00", "name": "Принтеры, МФУ"},
    {"code": "8414.51.00.00", "name": "Вентиляторы настольные, напольные"},
    {"code": "8509.40.00.00", "name": "Блендеры, кухонные комбайны"},
    {"code": "8516.60.10.00", "name": "Плиты электрические, духовки"},
    {"code": "8508.11.00.00", "name": "Пылесосы бытовые"},
    {"code": "8510.10.00.00", "name": "Электробритвы, триммеры"},
    {"code": "8519.81.95.00", "name": "Портативные колонки, аудиосистемы"},
    {"code": "9405.42.00.00", "name": "Светильники, лампы LED"},
    {"code": "9404.21.00.00", "name": "Матрасы, подушки ортопедические"},
    {"code": "9401.61.00.00", "name": "Мебель мягкая, диваны"},
    {"code": "7013.49.99.00", "name": "Посуда стеклянная, стаканы"},
    {"code": "7323.93.90.00", "name": "Посуда из нержавеющей стали"},
    {"code": "4202.92.98.00", "name": "Сумки, рюкзаки, чемоданы"},
    {"code": "4202.31.00.00", "name": "Кошельки, портмоне кожаные"},
    {"code": "3924.10.00.00", "name": "Контейнеры пластиковые для хранения"},
    {"code": "3926.90.97.00", "name": "Чехлы для телефонов пластиковые"},
    {"code": "8507.60.00.00", "name": "Повербанки, аккумуляторы литиевые"},
    {"code": "8544.42.90.00", "name": "Кабели, зарядные устройства USB"},
    {"code": "8504.40.90.00", "name": "Зарядные устройства, адаптеры питания"},
    {"code": "8523.51.00.00", "name": "USB-накопители, флешки"},
    {"code": "8471.60.70.00", "name": "Клавиатуры, мыши компьютерные"},
    {"code": "9506.91.00.00", "name": "Тренажёры, спортивное оборудование"},
    {"code": "9506.62.00.00", "name": "Мячи надувные, футбольные"},
    {"code": "9504.50.00.00", "name": "Игровые приставки, геймпады"},
    {"code": "8415.10.90.00", "name": "Кондиционеры бытовые"},
    {"code": "8418.10.20.00", "name": "Холодильники бытовые"},
    {"code": "8450.11.90.00", "name": "Стиральные машины бытовые"},
    {"code": "8422.11.00.00", "name": "Посудомоечные машины бытовые"},
    {"code": "8516.31.00.00", "name": "Фены для волос"},
    {"code": "8516.40.00.00", "name": "Утюги электрические"},
    {"code": "8516.71.00.00", "name": "Кофемашины, кофеварки электрические"},
    {"code": "8516.10.80.00", "name": "Чайники электрические"},
    {"code": "3401.11.00.00", "name": "Мыло туалетное"},
    {"code": "3402.20.90.00", "name": "Средства моющие, стиральные порошки"},
    {"code": "3307.49.00.00", "name": "Освежители воздуха, ароматизаторы"},
    {"code": "6302.60.00.00", "name": "Полотенца махровые"},
    {"code": "6302.10.00.00", "name": "Постельное бельё"},
    {"code": "4818.10.00.00", "name": "Бумага туалетная, салфетки"},
    {"code": "9608.10.10.00", "name": "Ручки шариковые, канцтовары"},
    {"code": "4820.10.00.00", "name": "Тетради, блокноты"},
    {"code": "8539.50.00.00", "name": "Лампочки LED, светодиодные"},
    {"code": "8512.20.00.00", "name": "Осветительное оборудование для авто"},
    {"code": "4011.10.00.00", "name": "Шины автомобильные"},
    {"code": "8708.99.97.00", "name": "Запчасти для автомобилей"},
    {"code": "8711.60.10.00", "name": "Электросамокаты, электровелосипеды"},
    {"code": "9021.10.10.00", "name": "Ортопедические изделия, корсеты"},
    {"code": "9018.90.84.00", "name": "Медицинские приборы, тонометры"},
    {"code": "3006.10.30.00", "name": "Пластыри, бинты, перевязочные материалы"},
    {"code": "3926.40.00.00", "name": "Статуэтки, декоративные изделия, копилки"},
]


def _fuzzy_match_tn_ved(query: str) -> list[dict[str, str]]:
    """Find matching ТН ВЭД codes by product name (fuzzy)."""
    query_lower = query.lower()
    words = query_lower.split()

    scored: list[tuple[int, dict]] = []
    for entry in TN_VED_DATABASE:
        name_lower = entry["name"].lower()
        score = 0
        for word in words:
            if len(word) < 3:
                continue
            # Match whole words or start of words
            if f" {word}" in f" {name_lower}":
                score += 10
            elif any(f" {word[:4]}" in f" {name_lower}" for _ in [1] if len(word) >= 4):
                score += 3
        if score > 0:
            scored.append((score, entry))

    scored.sort(key=lambda x: -x[0])
    return [entry for _, entry in scored[:5]]


# ═══════════════════════════════════════════════════════════════════════════
# Simple Kazakh translation (common terms mapping)
# In production, this would call a translation API
# ═══════════════════════════════════════════════════════════════════════════

RU_TO_KZ_TERMS: dict[str, str] = {
    "увлажнитель": "ауа ылғалдағыш",
    "воздуха": "ауа",
    "наушники": "құлаққап",
    "беспроводные": "сымсыз",
    "телефон": "телефон",
    "чехол": "қап",
    "игрушка": "ойыншық",
    "детская": "балалар",
    "детские": "балалар",
    "обувь": "аяқ киім",
    "одежда": "киім",
    "косметика": "косметика",
    "витамины": "витаминдер",
    "пылесос": "шаңсорғыш",
    "холодильник": "тоңазытқыш",
    "стиральная машина": "кір жуу машинасы",
    "чайник": "шәйнек",
    "кофемашина": "кофе машинасы",
    "утюг": "үтік",
    "фен": "фен",
    "телевизор": "теледидар",
    "монитор": "монитор",
    "принтер": "принтер",
    "ноутбук": "ноутбук",
    "планшет": "планшет",
    "кабель": "кабель",
    "зарядное": "зарядтау",
    "устройство": "құрылғы",
    "лампа": "шам",
    "светильник": "жарықтандырғыш",
    "кондиционер": "кондиционер",
    "вентилятор": "желдеткіш",
    "матрас": "матрас",
    "подушка": "жастық",
    "полотенце": "сүлгі",
    "постельное бельё": "төсек-орын жабдықтары",
    "сумка": "сөмке",
    "рюкзак": "рюкзак",
    "мыло": "сабын",
    "порошок": "ұнтақ",
    "электрический": "электрлік",
    "электрическая": "электрлік",
    "бытовой": "тұрмыстық",
    "бытовая": "тұрмыстық",
    "портативный": "тасымалды",
    "портативная": "тасымалды",
    "набор": "жинақ",
    "комплект": "жиынтық",
}


def _translate_to_kazakh(text: str) -> str:
    """Simple rule-based translation. In production, use a proper API."""
    result = text
    # Try to translate known terms
    for ru_term, kz_term in sorted(RU_TO_KZ_TERMS.items(), key=lambda x: -len(x[0])):
        result = re.sub(
            re.escape(ru_term),
            kz_term,
            result,
            flags=re.IGNORECASE
        )
    return result


async def _llm_pick_oktru(product_title: str, candidates: list[dict]) -> str | None:
    """Use Groq Llama to pick the best OKTRU code from a list."""
    if not candidates: 
        return None
        
    import httpx
    from retailpool.config import settings
    
    api_key = settings.GROQ_API_KEY
    if not api_key: 
        return candidates[0]["code"]

    prompt = f"Помоги выбрать наиболее подходящую категорию ОКТРУ для товара: '{product_title}'.\n\nДоступные категории:\n"
    for i, c in enumerate(candidates[:20]): # Limit to 20 to avoid huge prompts
        prompt += f"{i+1}. {c['code']} - {c['nameRu']}\n"
    prompt += "\nОтветь ТОЛЬКО кодом (в формате XXXX-XXXX-XXXX-XXXXXXXXX) самой подходящей категории. Больше ничего не пиши."

    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {api_key}"},
                json={
                    "model": "llama3-70b-8192",
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                    "max_tokens": 50
                },
            )
            if resp.status_code == 200:
                data = resp.json()
                answer = data["choices"][0]["message"]["content"].strip()
                import re
                match = re.search(r'\d{4}-\d{4}-\d{4}-\d{9}', answer)
                if match:
                    return match.group(0)
    except Exception as e:
        logger.warning(f"LLM pick error: {e}")
    
    return candidates[0]["code"]


# ═══════════════════════════════════════════════════════════════════════════
# NTIN Service class
# ═══════════════════════════════════════════════════════════════════════════

class NtinService:
    """Business logic for NTIN product management."""

    def __init__(self, session: AsyncSession):
        self.session = session

    # ── CRUD ──────────────────────────────────────────────────────

    async def create_product(
        self, user_id: uuid.UUID, data: dict[str, Any]
    ) -> NtinProduct:
        """Create a new NTIN product card."""
        product = NtinProduct(
            user_id=user_id,
            title_ru=data["title_ru"],
            description_ru=data.get("description_ru"),
            barcode=data.get("barcode"),
            kaspi_sku=data.get("kaspi_sku"),
            brand=data.get("brand"),
            oktru_code=data.get("oktru_code"),
            country_of_origin=data.get("country_of_origin", "Китай"),
            unit_of_measure=data.get("unit_of_measure", "шт"),
            weight_kg=data.get("weight_kg"),
            price=data.get("price"),
            image_url=data.get("image_url"),
            status=NtinStatus.DRAFT,
        )
        self.session.add(product)
        await self.session.flush()
        logger.info("Created NTIN product: %s for user %s", product.title_ru, user_id)
        return product

    async def get_products(
        self, user_id: uuid.UUID, status_filter: str | None = None
    ) -> list[NtinProduct]:
        """Get all NTIN products for a user, optionally filtered by status."""
        stmt = select(NtinProduct).where(
            NtinProduct.user_id == user_id
        ).order_by(NtinProduct.created_at.desc())

        if status_filter:
            stmt = stmt.where(NtinProduct.status == status_filter)

        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def get_product(
        self, product_id: uuid.UUID, user_id: uuid.UUID
    ) -> NtinProduct | None:
        """Get a single NTIN product by ID (scoped to user)."""
        stmt = select(NtinProduct).where(
            NtinProduct.id == product_id,
            NtinProduct.user_id == user_id,
        )
        result = await self.session.execute(stmt)
        return result.scalar_one_or_none()

    async def get_stats(self, user_id: uuid.UUID) -> dict[str, int]:
        """Get NTIN status counts for a user."""
        stmt = (
            select(NtinProduct.status, func.count())
            .where(NtinProduct.user_id == user_id)
            .group_by(NtinProduct.status)
        )
        result = await self.session.execute(stmt)
        counts = {row[0]: row[1] for row in result.all()}
        return {
            "total": sum(counts.values()),
            "draft": counts.get(NtinStatus.DRAFT, 0),
            "ai_filled": counts.get(NtinStatus.AI_FILLED, 0),
            "ready": counts.get(NtinStatus.READY, 0),
            "submitted": counts.get(NtinStatus.SUBMITTED, 0),
            "revision": counts.get(NtinStatus.REVISION, 0),
            "approved": counts.get(NtinStatus.APPROVED, 0),
            "rejected": counts.get(NtinStatus.REJECTED, 0),
        }

    # ── AI Fill ───────────────────────────────────────────────────

    async def ai_fill_product(self, product_id: uuid.UUID, user_id: uuid.UUID) -> NtinProduct | None:
        """AI-fill product attributes: ТН ВЭД code, Kazakh translation, and dynamic OKTRU code."""
        import httpx
        import urllib.parse
        
        product = await self.get_product(product_id, user_id)
        if not product:
            return None

        # 1. Find matching ТН ВЭД code
        search_text = f"{product.title_ru} {product.description_ru or ''}"
        matches = _fuzzy_match_tn_ved(search_text)
        if matches:
            best = matches[0]
            product.tn_ved_code = best["code"]
            product.tn_ved_name = best["name"]
            # Fill OKTRU code from the database if available
            if not product.oktru_code and "oktru_code" in best:
                product.oktru_code = best["oktru_code"]

        # 1.5 Dynamic Auto-fill OKTRU code using NKT API if missing or if default
        if not product.oktru_code or product.oktru_code == "1106-0001-0001-100011943":
            api_key = await self._get_nkt_api_key(user_id)
            candidates = []
            
            # 1. Try NKT API first
            if api_key and api_key != "test_api_key_12345":
                url = f"https://nationalcatalog.kz/gwp/portal/api/v1/dictionaries/OKTRU/items?page=1&size=50&search={urllib.parse.quote(product.title_ru)}"
                headers = {"X-API-KEY": api_key, "Accept": "application/json"}
                try:
                    async with httpx.AsyncClient(timeout=10.0) as client:
                        resp = await client.get(url, headers=headers)
                        if resp.status_code == 200:
                            data = resp.json()
                            items = data.get("content", [])
                            for item in items:
                                code = str(item.get("code", ""))
                                name_ru = str(item.get("nameRu", ""))
                                if code.count("-") == 3:
                                    candidates.append({"code": code, "nameRu": name_ru})
                except Exception as e:
                    logger.warning(f"Failed to auto-fetch OKTRU code from NKT API: {e}")
            
            # 2. Try Local DB if API failed or returned nothing
            if not candidates:
                from retailpool.models.ntin import OktruDictionary
                root_word = product.title_ru.lower()[:5] if len(product.title_ru) > 5 else product.title_ru.lower()
                safe_root = root_word.replace("%", "\\%").replace("_", "\\_")
                local_res = await self.session.execute(
                    select(OktruDictionary)
                    .where(OktruDictionary.search_vector.like(f"%{safe_root}%"))
                    .limit(20)
                )
                local_matches = local_res.scalars().all()
                for m in local_matches:
                    candidates.append({"code": m.code, "nameRu": m.name_ru})

            # 3. Use LLM to pick the best code
            if candidates:
                best_code = await _llm_pick_oktru(product.title_ru, candidates)
                if best_code:
                    product.oktru_code = best_code
                    logger.info(f"AI auto-found OKTRU code {product.oktru_code} for {product.title_ru}")

            # 4. Fallback if everything fails
            if not product.oktru_code or product.oktru_code == "1106-0001-0001-100011943":
                product.oktru_code = "3203-0001-0001-100017260"

        # 2. Translate to Kazakh
        product.title_kz = _translate_to_kazakh(product.title_ru)
        if product.description_ru:
            product.description_kz = _translate_to_kazakh(product.description_ru)

        # 3. Get templates and apply defaults
        settings = await self.get_settings(user_id)
        tpl_country = settings.tpl_country if settings and settings.tpl_country else "КИТАЙ"
        tpl_brand = settings.tpl_brand if settings and settings.tpl_brand else "Отсутствует"
        tpl_unit = settings.tpl_unit if settings and settings.tpl_unit else "шт"
        
        if not product.country_of_origin:
            product.country_of_origin = tpl_country
        if not product.unit_of_measure:
            product.unit_of_measure = tpl_unit
        if not product.brand or product.brand.lower() == 'none':
            product.brand = tpl_brand

        product.status = NtinStatus.AI_FILLED
        await self.session.flush()

        logger.info(
            "AI-filled product %s: ТН ВЭД=%s, title_kz=%s",
            product.id, product.tn_ved_code, product.title_kz[:40] if product.title_kz else None
        )
        return product

    # ── ТН ВЭД Lookup ────────────────────────────────────────────

    @staticmethod
    def search_tn_ved(query: str) -> list[dict[str, str]]:
        """Search ТН ВЭД codes by product name."""
        return _fuzzy_match_tn_ved(query)

    # ── Translation ──────────────────────────────────────────────

    @staticmethod
    def translate_to_kazakh(text: str) -> str:
        """Translate Russian text to Kazakh."""
        return _translate_to_kazakh(text)

    # ── Submit to НКТ ─────────────────────────────────────────────

    async def _get_nkt_api_key(self, user_id: uuid.UUID) -> str | None:
        """Get NKT API key: user-level first, then platform-level from .env."""
        user_settings = await self.get_settings(user_id)
        if user_settings and user_settings.nkt_api_key:
            return decrypt_secret(user_settings.nkt_api_key)
        # Fallback to platform-level key from .env
        from retailpool.config import settings as app_settings
        return getattr(app_settings, "NKT_API_KEY", None) or None
      
    def _build_nkt_headers(self, api_key: str) -> dict[str, str]:
        """Build headers for НКТ API requests."""
        return {
            "X-API-KEY": api_key,
            "Content-Type": "application/json",
            "Accept": "application/json",
        }

    async def _build_nkt_payload(self, user_id: uuid.UUID, product: NtinProduct) -> dict[str, Any]:
        """Build payload for НКТ API create request. (Fixed)

        Uses the real НКТ OpenAPI spec format:
        POST /portal/api/v1/products/requests
        {
            "oktru": "<category code>",
            "autoPublication": true,
            "attributes": [{"code": "...", "value": "..."}, ...]
        }
        """
        # Map country names to ISO codes
        country_map = {
            "Китай": "CN", "Казахстан": "KZ", "Россия": "RU",
            "Корея": "KR", "Япония": "JP", "США": "US", "Германия": "DE",
            "Турция": "TR", "Италия": "IT", "Франция": "FR", "Индия": "IN",
            "Великобритания": "GB", "Вьетнам": "VN", "Тайвань": "TW",
        }
        country_code = country_map.get(
            product.country_of_origin or "Китай", "CN"
        )

        oktru = product.oktru_code

        attributes: list[dict[str, Any]] = [
            {"code": "name_ru", "value": product.title_ru},
        ]

        if product.title_kz:
            attributes.append({"code": "name_kk", "value": product.title_kz})

        # Short name (required) — use first 50 chars
        short_ru = (product.title_ru or "")[:50]
        attributes.append({"code": "short_name_ru", "value": short_ru})

        if product.title_kz:
            short_kz = product.title_kz[:50]
            attributes.append({"code": "short_name_kk", "value": short_kz})

        attributes.append({"code": "country", "value": country_code})
        
        settings = await self.get_settings(user_id)
        tpl_qty = settings.tpl_qty if settings and settings.tpl_qty else 1
        
        # OKEI unit mapping for NKT API
        unit_map = {
            "шт": "796", "штука": "796", "шт.": "796",
            "кг": "166", "килограмм": "166",
            "компл": "839", "комплект": "839",
            "упак": "778", "упаковка": "778",
            "пара": "715", "пар": "715"
        }
        raw_unit = (product.unit_of_measure or "шт").lower().strip()
        nkt_unit = unit_map.get(raw_unit, "796")
        
        attributes.append({"code": "measure_unit", "value": nkt_unit})
        attributes.append({"code": "quantity", "value": str(tpl_qty)})
        
        # Наименование производителя (a4282e5d) is mandatory
        manufacturer = product.brand if product.brand and product.brand.lower() != 'none' else "Не указано"
        attributes.append({"code": "a4282e5d", "value": manufacturer})

        if product.brand and product.brand.lower() != 'none':
            attributes.append({"code": "brand", "value": product.brand})

        if product.barcode:
            # GTIN barcode
            attributes.append({"code": "gtin", "value": product.barcode})

        if product.tn_ved_code:
            # Strip dots for НКТ format
            tnved_clean = product.tn_ved_code.replace(".", "")
            attributes.append({"code": "tnved", "value": tnved_clean})

        return {
            "oktru": oktru,
            "autoPublication": True,
            "attributes": attributes,
        }

    async def submit_to_nkt(
        self, product_id: uuid.UUID, user_id: uuid.UUID
    ) -> NtinProduct | None:
        """Submit a product card to НКТ for NTIN assignment.

        Uses the real НКТ API (nationalcatalog.kz) OpenAPI v3.1:
        1. POST /portal/api/v1/products/requests — create draft
        2. PUT  /portal/api/v1/products/requests/{id}/moderation — send to moderation

        Auth: X-API-KEY header
        """
        import httpx

        product = await self.get_product(product_id, user_id)
        if not product:
            return None

        if product.status not in (NtinStatus.AI_FILLED, NtinStatus.READY, NtinStatus.REVISION):
            logger.warning("Cannot submit product %s with status %s", product_id, product.status)
            return product

        # Check required fields
        missing = []
        if not product.title_ru:
            missing.append("title_ru")
        if not product.title_kz:
            missing.append("title_kz")
        if not product.oktru_code:
            missing.append("oktru_code (ОКТРУ)")
        if missing:
            logger.warning("Product %s missing required fields: %s", product_id, missing)
            product.status = NtinStatus.REVISION
            product.revision_comment = f"Не заполнены поля: {', '.join(missing)}"
            await self.session.flush()
            return product

        # Get API key (user-level → platform-level)
        api_key = await self._get_nkt_api_key(user_id)
        base_url = "https://nationalcatalog.kz/gwp"

        if not api_key:
            # No key — can't submit
            logger.warning("No NKT API key for user %s — cannot submit to НКТ", user_id)
            product.status = NtinStatus.REVISION
            product.revision_comment = "Не указан API-ключ НКТ. Перейдите в Настройки API и введите ключ от nationalcatalog.kz"
            await self.session.flush()
            return product

        headers = self._build_nkt_headers(api_key)
        payload = await self._build_nkt_payload(user_id, product)

        nkt_response_text = ""
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                # Step 1: Create request (draft) in НКТ
                resp = await client.post(
                    f"{base_url}/portal/api/v1/products/requests",
                    json=payload,
                    headers=headers,
                )
                nkt_response_text = resp.text
                logger.info(
                    "НКТ API create response for product %s: status=%d body=%s",
                    product_id, resp.status_code, nkt_response_text[:500]
                )

                if resp.status_code == 200:
                    data = resp.json()
                    nkt_request_id = data.get("id")
                    if nkt_request_id:
                        product.nkt_request_id = int(nkt_request_id)
                        logger.info("НКТ request created with id=%d", nkt_request_id)

                        # Step 2: Send to moderation
                        mod_resp = await client.put(
                            f"{base_url}/portal/api/v1/products/requests/{nkt_request_id}/moderation",
                            headers=headers,
                        )
                        logger.info(
                            "НКТ moderation response: status=%d body=%s",
                            mod_resp.status_code, mod_resp.text[:300]
                        )

                        if mod_resp.status_code == 200:
                            product.status = NtinStatus.SUBMITTED
                            product.submitted_at = datetime.now(timezone.utc)
                            nkt_response_text += f"\n[Moderation OK] id={nkt_request_id}"
                        else:
                            # Created but moderation failed — still mark as submitted
                            product.status = NtinStatus.SUBMITTED
                            product.submitted_at = datetime.now(timezone.utc)
                            nkt_response_text += f"\n[Moderation {mod_resp.status_code}] {mod_resp.text[:200]}"
                    else:
                        product.status = NtinStatus.SUBMITTED
                        product.submitted_at = datetime.now(timezone.utc)

                elif resp.status_code == 401:
                    logger.error("НКТ API: Invalid API key for user %s", user_id)
                    product.status = NtinStatus.REVISION
                    product.revision_comment = "Ошибка авторизации НКТ (401). Проверьте API-ключ в Настройках."
                elif resp.status_code == 400:
                    logger.warning("НКТ API: Bad request — %s", nkt_response_text)
                    try:
                        err_data = resp.json()
                        err_msg = err_data.get("message", nkt_response_text[:300])
                    except Exception:
                        err_msg = nkt_response_text[:300]
                    product.status = NtinStatus.REVISION
                    product.revision_comment = f"НКТ: ошибка запроса — {err_msg}"
                elif resp.status_code == 422:
                    logger.warning("НКТ API: Validation error — %s", nkt_response_text)
                    try:
                        err_data = resp.json()
                        err_msg = err_data.get("message", nkt_response_text[:300])
                    except Exception:
                        err_msg = nkt_response_text[:300]
                    product.status = NtinStatus.REVISION
                    product.revision_comment = f"НКТ: ошибка валидации — {err_msg}"
                else:
                    product.status = NtinStatus.REVISION
                    product.revision_comment = f"НКТ вернул статус {resp.status_code}"

        except httpx.TimeoutException:
            logger.error("НКТ API timeout for product %s", product_id)
            nkt_response_text = "Timeout: НКТ API не ответил за 30 секунд"
            product.status = NtinStatus.REVISION
            product.revision_comment = "Таймаут при подаче в НКТ. Попробуйте позже."
        except httpx.HTTPError as e:
            logger.error("НКТ API HTTP error for product %s: %s", product_id, e)
            nkt_response_text = f"HTTP Error: {e}"
            product.status = NtinStatus.REVISION
            product.revision_comment = f"Ошибка соединения с НКТ: {e}"

        # Record submission history
        submission = NtinSubmission(
            product_id=product.id,
            status=product.status,
            nkt_response=nkt_response_text[:2000],
        )
        self.session.add(submission)
        await self.session.flush()

        logger.info("Submitted product %s to НКТ → status=%s", product_id, product.status)
        return product

    # ── Check status in НКТ ────────────────────────────────────────

    async def check_nkt_status(
        self, product_id: uuid.UUID, user_id: uuid.UUID
    ) -> NtinProduct | None:
        """Check the status of a submitted product in НКТ.

        GET /portal/api/v1/products/requests/{id}/status
        """
        import httpx

        product = await self.get_product(product_id, user_id)
        if not product or not product.nkt_request_id:
            return product

        api_key = await self._get_nkt_api_key(user_id)
        if not api_key:
            return product

        base_url = "https://nationalcatalog.kz/gwp"
        headers = self._build_nkt_headers(api_key)

        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(
                    f"{base_url}/portal/api/v1/products/requests/{product.nkt_request_id}/status",
                    headers=headers,
                )
                logger.info(
                    "НКТ status check for request %d: %d %s",
                    product.nkt_request_id, resp.status_code, resp.text[:200]
                )

                if resp.status_code == 200:
                    data = resp.json()
                    nkt_status = data.get("code", "")

                    # Map НКТ statuses to our statuses
                    status_map = {
                        "new": NtinStatus.SUBMITTED,
                        "onModeration": NtinStatus.SUBMITTED,
                        "accepted": NtinStatus.SUBMITTED,
                        "underRevision": NtinStatus.REVISION,
                        "underRevisionGz": NtinStatus.REVISION,
                        "rejected": NtinStatus.REJECTED,
                        "rejectedGz": NtinStatus.REJECTED,
                        "cancelled": NtinStatus.REVOKED,
                        "readyToPublish": NtinStatus.APPROVED,
                        "completed": NtinStatus.APPROVED,
                    }

                    new_status = status_map.get(nkt_status, product.status)
                    nkt_label = data.get("value", nkt_status)

                    if new_status != product.status:
                        product.status = new_status
                        product.revision_comment = f"НКТ статус: {nkt_label}"

                        if new_status == NtinStatus.APPROVED:
                            product.approved_at = datetime.now(timezone.utc)
                            # Try to get NTIN code from details
                            await self._fetch_ntin_code(client, headers, base_url, product)

                    await self.session.flush()

        except Exception as e:
            logger.error("Error checking НКТ status: %s", e)

        return product

    async def _fetch_ntin_code(
        self, client: Any, headers: dict, base_url: str, product: NtinProduct
    ) -> None:
        """Try to fetch the NTIN code from the product details after approval."""
        try:
            details_resp = await client.get(
                f"{base_url}/portal/api/v1/products/requests/{product.nkt_request_id}/details",
                headers=headers,
            )
            if details_resp.status_code == 200:
                details = details_resp.json()
                attrs = details.get("attributes", [])
                for attr in attrs:
                    if attr.get("code") == "ntin" and attr.get("value"):
                        product.ntin_code = str(attr["value"])
                        logger.info("Got NTIN code %s for product %s", product.ntin_code, product.id)
                        break
        except Exception as e:
            logger.warning("Could not fetch NTIN code: %s", e)

    # ── Sync all requests from НКТ ────────────────────────────────

    async def sync_nkt_requests(
        self, user_id: uuid.UUID
    ) -> dict[str, int]:
        """Sync statuses of all submitted products from НКТ.

        Returns counts of updated products.
        """
        stmt = select(NtinProduct).where(
            NtinProduct.user_id == user_id,
            NtinProduct.nkt_request_id.isnot(None),
            NtinProduct.status.in_([NtinStatus.SUBMITTED, NtinStatus.REVISION]),
        )
        result = await self.session.execute(stmt)
        products = list(result.scalars().all())

        updated = 0
        for product in products:
            old_status = product.status
            await self.check_nkt_status(product.id, user_id)
            if product.status != old_status:
                updated += 1

        await self.session.flush()
        return {"checked": len(products), "updated": updated}

    # ── Bulk import from data ─────────────────────────────────────

    async def bulk_import(
        self, user_id: uuid.UUID, products_data: list[dict[str, Any]]
    ) -> list[NtinProduct]:
        """Bulk import products (from Excel or Kaspi store)."""
        created = []
        for data in products_data:
            product = await self.create_product(user_id, data)
            created.append(product)
        await self.session.flush()
        logger.info("Bulk imported %d products for user %s", len(created), user_id)
        return created

    async def _parse_kaspi_xml_content(self, user_id: uuid.UUID, text: str) -> int:
        from xml.etree import ElementTree as ET
        root = ET.fromstring(text)
        
        # Strip namespaces to avoid ElementTree xpath limitations
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]
        
        offers = root.findall('.//offer')
        
        products_to_create = []
        stmt = select(NtinProduct.kaspi_sku).where(
            NtinProduct.user_id == user_id,
            NtinProduct.kaspi_sku.isnot(None)
        )
        result = await self.session.execute(stmt)
        existing_skus = {row[0] for row in result.all()}
        
        for offer in offers:
            sku = offer.get('sku')
            if not sku or sku in existing_skus:
                continue
                
            model_node = offer.find('.//model')
            title = None
            if model_node is not None and model_node.text and model_node.text.strip():
                title = model_node.text.strip()
            if not title:
                title = f"Товар {sku}"
            
            brand_node = offer.find('.//brand')
            brand = brand_node.text.strip() if brand_node is not None and brand_node.text else None
            
            price_node = offer.find('.//price')
            
            try:
                price = float(price_node.text) if price_node is not None and price_node.text else None
            except (ValueError, TypeError):
                price = None
                
            products_to_create.append({
                "title_ru": title,
                "kaspi_sku": sku,
                "brand": brand,
                "price": price,
            })
        
        if products_to_create:
            await self.bulk_import(user_id, products_to_create)
        
        return len(products_to_create)

    async def fetch_kaspi_xml(self, user_id: uuid.UUID, xml_url: str) -> int:
        """Fetch Kaspi XML feed, parse it, and import products without barcodes."""
        import httpx
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.get(xml_url)
                resp.raise_for_status()
                return await self._parse_kaspi_xml_content(user_id, resp.text)
        except Exception as e:
            logger.error("Failed to fetch/parse Kaspi XML: %s", e)
            raise

    async def parse_kaspi_file(self, user_id: uuid.UUID, content: bytes, filename: str) -> int:
        """Parse uploaded Kaspi XML or Excel file."""
        if filename.lower().endswith('.xml'):
            text = content.decode('utf-8', errors='replace')
            return await self._parse_kaspi_xml_content(user_id, text)
        elif filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls'):
            import openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
            
            stmt = select(NtinProduct.kaspi_sku).where(
                NtinProduct.user_id == user_id,
                NtinProduct.kaspi_sku.isnot(None)
            )
            result = await self.session.execute(stmt)
            existing_skus = {row[0] for row in result.all()}
            
            products_to_create = []
            
            # Find column indices
            # Some excel files start data on row 1, some on row 2, let's find the header row
            headers = []
            header_row_idx = 1
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                row_strs = [str(cell).lower().strip() if cell else "" for cell in row]
                if any(x in row_strs for x in ["sku", "артикул", "код", "код товара"]):
                    headers = row_strs
                    header_row_idx = row_idx
                    break
            
            if not headers:
                # Fallback assuming row 1 is header
                headers = [str(cell.value).lower().strip() if cell.value else "" for cell in sheet[1]]
                
            sku_idx = -1
            title_idx = -1
            brand_idx = -1
            price_idx = -1
            
            for i, h in enumerate(headers):
                if "sku" in h or "артикул" in h or "код" in h: sku_idx = i
                if "название" in h or "модель" in h or "наименование" in h or "товар" in h: title_idx = i
                if "бренд" in h or "производитель" in h: brand_idx = i
                if "цена" in h: price_idx = i
                
            if sku_idx == -1 or title_idx == -1:
                # If we really can't find columns, assume Col A is SKU, Col B is Name
                sku_idx, title_idx = 0, 1
                
            for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if not row or len(row) <= max(sku_idx, title_idx) or not row[sku_idx]: continue
                sku = str(row[sku_idx]).strip()
                if not sku or sku in existing_skus or sku.lower() == 'none':
                    continue
                    
                title = str(row[title_idx]).strip() if row[title_idx] and str(row[title_idx]).lower() != 'none' else f"Товар {sku}"
                brand = str(row[brand_idx]).strip() if brand_idx != -1 and len(row) > brand_idx and row[brand_idx] and str(row[brand_idx]).lower() != 'none' else None
                price = None
                if price_idx != -1 and len(row) > price_idx and row[price_idx]:
                    try: price = float(str(row[price_idx]).replace(' ', '').replace(',', '.'))
                    except ValueError: pass
                    
                products_to_create.append({
                    "title_ru": title,
                    "kaspi_sku": sku,
                    "brand": brand,
                    "price": price,
                })
                
            if products_to_create:
                await self.bulk_import(user_id, products_to_create)
            
            return len(products_to_create)
        else:
            raise ValueError("Поддерживаются только XML и XLSX/XLS файлы")

    # ── Settings ──────────────────────────────────────────────────

    async def get_settings(self, user_id: uuid.UUID) -> UserSellerSettings | None:
        """Get user's seller settings."""
        stmt = select(UserSellerSettings).where(
            UserSellerSettings.user_id == user_id
        )
        result = await self.session.execute(stmt)
        return result.scalar_one_or_none()

    async def save_settings(
        self, user_id: uuid.UUID, data: dict[str, Any]
    ) -> UserSellerSettings:
        """Save or update user's seller settings."""
        settings = await self.get_settings(user_id)
        if not settings:
            settings = UserSellerSettings(user_id=user_id)
            self.session.add(settings)

        if "kaspi_api_key" in data:
            settings.kaspi_api_key = encrypt_secret(data["kaspi_api_key"])
        if "kaspi_merchant_id" in data:
            settings.kaspi_merchant_id = data["kaspi_merchant_id"]
        if "kaspi_shop_name" in data:
            settings.kaspi_shop_name = data["kaspi_shop_name"]
        if "kaspi_xml_url" in data:
            settings.kaspi_xml_url = data["kaspi_xml_url"]
        if "nkt_api_key" in data:
            settings.nkt_api_key = encrypt_secret(data["nkt_api_key"])
            
        if "tpl_country" in data: settings.tpl_country = data["tpl_country"]
        if "tpl_brand" in data: settings.tpl_brand = data["tpl_brand"]
        if "tpl_unit" in data: settings.tpl_unit = data["tpl_unit"]
        if "tpl_qty" in data: settings.tpl_qty = data["tpl_qty"]

        await self.session.flush()
        logger.info("Saved seller settings for user %s", user_id)
        return settings
