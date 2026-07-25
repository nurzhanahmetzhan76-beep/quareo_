import io
import openpyxl

def test():
    # Create an empty workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    print("max_column:", ws.max_column)
    print("max_row:", ws.max_row)

    # Add headers
    ws.append(["SKU", "Price", "Name"])
    ws.append(["12345", "1000", "Product"])

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            headers[str(v).strip().lower()] = c
    print("Headers:", headers)

    if "sku" not in headers or "price" not in headers:
        print("400 Error")
        return

    sku_col = headers["sku"]
    price_col = headers["price"]

    name_col = None
    for h_key, h_col in headers.items():
        if any(p in h_key for p in ["model", "name", "название", "наименовани", "title"]):
            name_col = h_col
            break

    if not name_col:
        name_col = sku_col

    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, sku_col).value
        if sku is None:
            continue
        sku = str(sku).strip()

        name = ws.cell(r, name_col).value
        name = str(name).strip() if name else f"Товар {sku}"

        price_val = ws.cell(r, price_col).value
        try:
            price = float(price_val) if price_val is not None else 0
        except ValueError:
            price = 0

        print(f"Row {r}: sku={sku}, name={name}, price={price}")

if __name__ == "__main__":
    try:
        test()
    except Exception as e:
        import traceback
        traceback.print_exc()
