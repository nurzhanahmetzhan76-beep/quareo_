import openpyxl

try:
    wb = openpyxl.load_workbook('Shablon-prais-lista.xlsx', data_only=True)
    
    with open('headers_output.txt', 'w', encoding='utf-8') as f:
        f.write('Sheets: ' + ','.join(wb.sheetnames) + '\n\n')
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
            row2 = [str(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)]
            
            f.write(f'--- {sheet_name} ---\n')
            f.write(','.join(headers) + '\n')
            f.write(','.join(row2) + '\n\n')
except Exception as e:
    with open('headers_output.txt', 'w', encoding='utf-8') as f:
        f.write('Error: ' + str(e))
